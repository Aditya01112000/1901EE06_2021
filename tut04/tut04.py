import os.path
import csv
from openpyxl import Workbook
from openpyxl import load_workbook

def output_by_subject():
    direc="output_by_subject"
     # check if file is not already exist ....if exist then updating the same folder
    if os.path.exists(direc)==False:
        os.mkdir(direc) 
    # opening input file
    with open("regtable_old.csv","r") as f:
        reader=csv.reader(f)
        temp=[]
        for words in reader:
            subject=""
            list1=[words[0],words[1],words[3],words[8]]
            if words[0]=="rollno":
                temp=list1.copy()
            else:
                subject=words[3]
                check=os.path.join(direc,subject+".xlsx")
                if os.path.isfile(check) is True:
                    wb = load_workbook(check)
                    sheet = wb.active
                    sheet.append(list1)
                    wb.save(check)
                else:
                    wb = Workbook()
                    sheet = wb.active
                    sheet.append(temp)
                    sheet.append(list1)     
                    wb.save(check)
    return

def output_individual_roll():
    direc="output_individual_roll"
    # check if file is not already exist ....if exist then updating the same folder
    if os.path.exists(direc)==False:
        os.mkdir(direc) 
    # opening input file
    with open('regtable_old.csv','r') as f:
        reader =csv.reader(f)
        temp=[]
        for words in reader:
            roll_no=""
            list1=[words[0],words[1],words[3],words[8]]
            if words[0]=="rollno":
                temp=list1.copy()
            else:
                roll_no=words[0]
                check=os.path.join(direc,roll_no+".xlsx")
                #checking if file already exist or not
                if os.path.isfile(check) is True:
                    wb = load_workbook(check)
                    sheet = wb.active
                    sheet.append(list1)
                    wb.save(check)
                else:
                    wb = Workbook()
                    sheet = wb.active
                    sheet.append(temp)
                    sheet.append(list1)     
                    wb.save(check)
    return

output_individual_roll()
output_by_subject()