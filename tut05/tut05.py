from os import name
import os.path
import csv
from openpyxl import Workbook
from openpyxl import load_workbook


d={}
subject_credit={}
direc="output"
grade_equ={
    "AA":10,
    "AB":9,
    "BB":8,
    "BC":7,
    "CC":6,
    "CD":5,
    "DD":4,
    "F":0,
    "I":0,
    "AA*":10,
    "AB*":9,
    "BB*":8,
    "BC*":7,
    "CC*":6,
    "CD*":5,
    "DD*":4,
    "F*":0,
    "I*":0
}

def mapping():
     with open('names-roll.csv','r') as f:
        reader =csv.reader(f)
        for words in reader:
            d[words[0]]=words[1]

def subject():
    with open('subjects_master.csv','r') as f:
        reader =csv.reader(f)
        for words in reader:
            subject_credit[words[0]]=[words[1],words[2]]


def generate_marksheet():
    # check if file is not already exist ....if exist then updating the same folder
    if os.path.exists(direc)==False:
        os.mkdir(direc)  
    with open('grades.csv','r') as f:
        reader =csv.reader(f)
        temp=[]
        roll_no=""
        sem=""
        for words in reader:
            roll_no=words[0]
            sem=words[1]
            if roll_no=="Roll":
                continue
            check=os.path.join(direc,roll_no+".xlsx")
            if os.path.isfile(check) is False:
                wb = Workbook()
                sheet = wb.active
                sheet.append(["Roll No.",roll_no])
                sheet.append(["Name of Student",d[roll_no]])  
                sheet.append(["Discipline",roll_no[4:6]]) 
                # sheet.append(["Semester No.",1,2,3,4,5,6,7,8])  
                sheet.title="Overall"
                wb.save(check)
            
            wb = load_workbook(check)
            sheet_name="Sem"+sem
            if sheet_name in wb.sheetnames:
                curr_sheet=wb[sheet_name]
                curr_sheet.append([curr_sheet.max_row,words[2],subject_credit[words[2]][0],subject_credit[words[2]][1],int(words[3]),words[5],words[4].strip()])
            else:
                wb.create_sheet(index=int(sem),title=sheet_name)
                curr_sheet=wb[sheet_name]
                curr_sheet.append(["Sl No.","Subject No,","Subject Name","L-T-P","Credit","Subject Type","Grade"])
                curr_sheet.append([1,words[2],subject_credit[words[2]][0],subject_credit[words[2]][1],int(words[3]),words[5],words[4].strip()])
            wb.save(check)

def spi():
        for jk in d.keys():
            roll_no=jk
            if(roll_no=="Roll"):
              continue
            check=os.path.join(direc,roll_no+".xlsx")
            wb = load_workbook(check)
            overall=wb["Overall"]
            semester_no=["Semester No."]
            semester_wise=["Semester wise Credit Taken"]
            spi=["SPI"]
            total_cre=["Total Credit"]
            cpi=["CPI"]
            total_temp=0
            cpi_temp=0
            for i in range(len(wb.sheetnames)):
                if wb.sheetnames[i]=="Overall":
                    continue
                curr_sheet=wb[wb.sheetnames[i]]
                total_credit=0
                avg=0;
                for j in range(1,curr_sheet.max_row):
                    curr_credit=curr_sheet.cell(row=j+1,column=5).value
                    curr_grade=curr_sheet.cell(row=j+1,column=7).value
                    total_credit=total_credit+curr_credit
                    avg=avg+(curr_credit*grade_equ[curr_grade])
                avg=avg/total_credit
                semester_no.append(int(wb.sheetnames[i][3:]))
                semester_wise.append(total_credit)
                spi.append(round(avg,2))
                cpi_temp=cpi_temp+(total_credit*avg)
                total_temp=total_temp+total_credit
                total_cre.append(total_temp)
                cpi.append(round(cpi_temp/total_temp,2))
            # print(semester_no)
            # print(semester_wise)
            # print(spi)
            # print(total_cre)
            # print(cpi)
            overall.append(semester_no)
            overall.append(semester_wise)
            overall.append(spi)
            overall.append(total_cre)
            overall.append(cpi)
            wb.save(check)

               
                  



    

mapping()
subject()
# generate_marksheet()
spi()